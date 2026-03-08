import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import app_engine
from app_engine import (
    build_session,
    cleanup_session_dir,
    is_supported_excel,
    progress_stats,
    sync_progress_to_source,
)


class AppEngineTests(unittest.TestCase):
    def test_supported_extension_check(self):
        self.assertTrue(is_supported_excel(Path("a.xlsx")))
        self.assertTrue(is_supported_excel(Path("a.xlsm")))
        self.assertTrue(is_supported_excel(Path("a.xls")))
        self.assertTrue(is_supported_excel(Path("a.xlsb")))
        self.assertTrue(is_supported_excel(Path("a.excel")))
        self.assertFalse(is_supported_excel(Path("a.csv")))

    def test_build_session_patrol_and_sync(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "patrol.xlsx"
            photo_path = base / "p1.jpg"
            photo_path.write_bytes(b"fake")

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            header = [
                "序号",
                "巡视日期",
                "问题道路",
                "具体位置",
                "问题表述",
                "问题类型",
                "巡 视 问 题",
                "处 置 情 况",
                "备  注",
                "维修性质",
                "计划完成时间",
                "实际完成时间",
                "单位",
                "维修工程量",
                "截止时间",
            ]

            for c, v in enumerate(header, start=1):
                ws.cell(1, c).value = v
                ws.cell(3, c).value = v

            ws.cell(2, 1).value = 1
            ws.cell(2, 3).value = "环湖西三路"
            ws.cell(2, 4).value = "近申港大道"
            ws.cell(2, 5).value = "绿化带有杂草"
            ws.cell(2, 7).value = str(photo_path)
            ws.cell(2, 15).value = "10"
            ws.cell(2, 16).value = "已填"
            ws.cell(2, 17).value = "2026-03-06 10:00:00"

            ws.cell(4, 1).value = 2
            ws.cell(4, 3).value = "环湖西三路"
            ws.cell(4, 4).value = "近申港大道"
            ws.cell(4, 5).value = "树木需要剥芽"
            ws.cell(4, 7).value = str(photo_path)
            ws.cell(4, 15).value = "10"

            wb.save(excel_path)
            wb.close()

            result = build_session(excel_path)
            self.assertEqual(result.mode, "patrol")
            self.assertEqual(result.total_records, 2)

            lines = result.paths.data_tsv.read_text(encoding="utf-8").splitlines()
            first = lines[1].split("\t")
            self.assertEqual(first[1], "环湖西三路")
            self.assertEqual(first[2], "近申港大道")

            done, total = progress_stats(result.paths.progress_tsv)
            self.assertEqual(total, 2)
            self.assertEqual(done, 1)

            # Simulate F10 writes for both rows
            result.paths.progress_tsv.write_text(
                "source_row\t状态\t提交时间\n"
                "2\t已填\t2026-03-06 11:00:00\n"
                "4\t已填\t2026-03-06 11:05:00\n",
                encoding="utf-8",
            )
            updated = sync_progress_to_source(result.paths.meta_json, result.paths.progress_tsv)
            self.assertEqual(updated, 2)

            ws2 = load_workbook(excel_path, data_only=True).active
            assert ws2 is not None
            self.assertEqual(ws2.cell(2, 16).value, "已填")
            self.assertEqual(ws2.cell(4, 16).value, "已填")

            cleanup_session_dir(result.paths.session_dir)

    def test_build_session_patrol_shifted_status_columns(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "patrol_shifted.xlsx"
            photo_path = base / "p1.jpg"
            photo_path.write_bytes(b"fake")

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            header = [
                "序号",
                "巡视日期",
                "问题道路",
                "具体位置",
                "问题表述",
                "问题类型",
                "巡 视 问 题",
                "处 置 情 况",
                "备  注",
                "维修性质",
                "计划完成时间",
                "实际完成时间",
                "单位",
                "维修工程量",
                "整改描述",
                "截止时间",
                "填报状态",
                "提交时间",
            ]

            for c, v in enumerate(header, start=1):
                ws.cell(1, c).value = v
                ws.cell(3, c).value = v

            ws.cell(2, 1).value = 1
            ws.cell(2, 3).value = "环湖西三路"
            ws.cell(2, 4).value = "近申港大道"
            ws.cell(2, 5).value = "绿化带有杂草"
            ws.cell(2, 7).value = str(photo_path)
            ws.cell(2, 15).value = "已除草"
            ws.cell(2, 16).value = "10"
            ws.cell(2, 17).value = "已填"
            ws.cell(2, 18).value = "2026-03-06 10:00:00"

            ws.cell(4, 1).value = 2
            ws.cell(4, 3).value = "环湖西三路"
            ws.cell(4, 4).value = "近申港大道"
            ws.cell(4, 5).value = "树木需要剥芽"
            ws.cell(4, 7).value = str(photo_path)
            ws.cell(4, 15).value = "已剥芽"
            ws.cell(4, 16).value = "10"

            wb.save(excel_path)
            wb.close()

            result = build_session(excel_path)
            done, total = progress_stats(result.paths.progress_tsv)
            self.assertEqual(total, 2)
            self.assertEqual(done, 1)

            result.paths.progress_tsv.write_text(
                "source_row\t状态\t提交时间\n"
                "2\t已填\t2026-03-06 11:00:00\n"
                "4\t已填\t2026-03-06 11:05:00\n",
                encoding="utf-8",
            )
            updated = sync_progress_to_source(result.paths.meta_json, result.paths.progress_tsv)
            self.assertEqual(updated, 2)

            ws2 = load_workbook(excel_path, data_only=True).active
            assert ws2 is not None
            self.assertEqual(ws2.cell(4, 16).value, "10")
            self.assertEqual(ws2.cell(4, 17).value, "已填")
            self.assertEqual(ws2.cell(4, 18).value, "2026-03-06 11:05:00")

            cleanup_session_dir(result.paths.session_dir)

    def test_build_session_patrol2_and_sync_independent_status(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "patrol_profile2.xlsx"
            photo_path = base / "p2.jpg"
            photo_path.write_bytes(b"fake")

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            header = [
                "序号",
                "巡视日期",
                "问题道路",
                "具体位置",
                "问题表述",
                "问题类型",
                "巡 视 问 题",
                "处 置 情 况",
                "备  注",
                "维修性质",
                "计划完成时间",
                "实际完成时间",
                "单位",
                "维修工程量",
                "整改描述",
                "截止时间",
                "填报状态",
                "提交时间",
            ]

            for c, v in enumerate(header, start=1):
                ws.cell(1, c).value = v

            ws.cell(2, 1).value = 1
            ws.cell(2, 3).value = "环湖西三路"
            ws.cell(2, 4).value = "近申港大道"
            ws.cell(2, 5).value = "绿化带有杂草"
            ws.cell(2, 8).value = str(photo_path)
            ws.cell(2, 15).value = "已完成整改描述"
            ws.cell(2, 16).value = "10"
            ws.cell(2, 17).value = "已填"  # 巡检填报1状态
            ws.cell(2, 18).value = "2026-03-06 10:00:00"

            wb.save(excel_path)
            wb.close()

            result = build_session(excel_path, fill_profile="patrol2")
            self.assertEqual(result.mode, "patrol")
            self.assertEqual(result.fill_profile, "patrol2")

            done, total = progress_stats(result.paths.progress_tsv)
            self.assertEqual(total, 1)
            self.assertEqual(done, 0)

            lines = result.paths.data_tsv.read_text(encoding="utf-8").splitlines()
            cols = lines[1].split("\t")
            self.assertEqual(cols[2], "环湖西三路")
            self.assertEqual(cols[5], "已完成整改描述")
            self.assertEqual(Path(cols[6]), photo_path)

            result.paths.progress_tsv.write_text(
                "source_row\t状态\t提交时间\n"
                "2\t已填\t2026-03-06 11:20:00\n",
                encoding="utf-8",
            )
            updated = sync_progress_to_source(result.paths.meta_json, result.paths.progress_tsv)
            self.assertEqual(updated, 1)

            ws2 = load_workbook(excel_path, data_only=True).active
            assert ws2 is not None
            self.assertEqual(ws2.cell(1, 19).value, "填报2状态")
            self.assertEqual(ws2.cell(1, 20).value, "填报2提交时间")
            self.assertEqual(ws2.cell(2, 17).value, "已填")
            self.assertEqual(ws2.cell(2, 18).value, "2026-03-06 10:00:00")
            self.assertEqual(ws2.cell(2, 19).value, "已填")
            self.assertEqual(ws2.cell(2, 20).value, "2026-03-06 11:20:00")

            cleanup_session_dir(result.paths.session_dir)

    def test_build_session_status_check_separated_by_profile(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "patrol_status_separated.xlsx"
            photo_path = base / "p4.jpg"
            photo_path.write_bytes(b"fake")

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            header = [
                "序号",
                "巡视日期",
                "问题道路",
                "具体位置",
                "问题表述",
                "问题类型",
                "巡 视 问 题",
                "处 置 情 况",
                "备  注",
                "维修性质",
                "计划完成时间",
                "实际完成时间",
                "单位",
                "维修工程量",
                "整改描述",
                "截止时间",
                "填报状态",
                "提交时间",
                "填报2状态",
                "填报2提交时间",
            ]

            for c, v in enumerate(header, start=1):
                ws.cell(1, c).value = v

            ws.cell(2, 1).value = 1
            ws.cell(2, 3).value = "环湖西三路"
            ws.cell(2, 4).value = "近申港大道"
            ws.cell(2, 5).value = "绿化带有杂草"
            ws.cell(2, 8).value = str(photo_path)
            ws.cell(2, 15).value = "已除草"
            ws.cell(2, 16).value = "10"

            # 仅填报2已完成，填报1未完成
            ws.cell(2, 19).value = "已填"
            ws.cell(2, 20).value = "2026-03-06 13:00:00"

            wb.save(excel_path)
            wb.close()

            patrol1 = build_session(excel_path, fill_profile="patrol1")
            done1, total1 = progress_stats(patrol1.paths.progress_tsv)
            self.assertEqual(total1, 1)
            self.assertEqual(done1, 0)

            patrol2 = build_session(excel_path, fill_profile="patrol2")
            done2, total2 = progress_stats(patrol2.paths.progress_tsv)
            self.assertEqual(total2, 1)
            self.assertEqual(done2, 1)

            cleanup_session_dir(patrol1.paths.session_dir)
            cleanup_session_dir(patrol2.paths.session_dir)

    def test_build_session_patrol2_accepts_chuliqingkuang_header(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "patrol_profile2_alias.xlsx"
            photo_path = base / "p3.jpg"
            photo_path.write_bytes(b"fake")

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            header = [
                "序号",
                "巡视日期",
                "问题道路",
                "具体位置",
                "问题表述",
                "问题类型",
                "巡 视 问 题",
                "处 理 情 况",
                "备  注",
                "维修性质",
                "计划完成时间",
                "实际完成时间",
                "单位",
                "维修工程量",
                "整改描述",
                "截止时间",
            ]

            for c, v in enumerate(header, start=1):
                ws.cell(1, c).value = v

            ws.cell(2, 1).value = 1
            ws.cell(2, 3).value = "古棕路"
            ws.cell(2, 4).value = "近申港大道"
            ws.cell(2, 5).value = "绿化带有杂草"
            ws.cell(2, 8).value = str(photo_path)
            ws.cell(2, 15).value = "已除草"
            ws.cell(2, 16).value = "10"

            wb.save(excel_path)
            wb.close()

            result = build_session(excel_path, fill_profile="patrol2")
            lines = result.paths.data_tsv.read_text(encoding="utf-8").splitlines()
            cols = lines[1].split("\t")

            self.assertEqual(cols[2], "古棕路")
            self.assertEqual(cols[5], "已除草")
            self.assertEqual(Path(cols[6]), photo_path)

            cleanup_session_dir(result.paths.session_dir)

    def test_build_session_simple(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "simple.xlsx"

            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(["问题地址", "问题路段", "截止时间", "问题类别", "问题描述", "上传照片", "处置方式", "状态", "提交时间"])
            ws.append(["A", "B", "12", "市政", "desc", "C:/tmp/x.jpg", "项目部安排", "", ""])
            ws.append(["A2", "B2", "24", "绿化", "desc2", "C:/tmp/y.jpg", "本班组执行", "已填", "2026-03-06 12:00:00"])
            wb.save(excel_path)
            wb.close()

            result = build_session(excel_path)
            self.assertEqual(result.mode, "simple")
            done, total = progress_stats(result.paths.progress_tsv)
            self.assertEqual(total, 2)
            self.assertEqual(done, 1)

            cleanup_session_dir(result.paths.session_dir)

    def test_patrol_image_strict_to_xunshi_column(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "patrol_images.xlsx"

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            header = [
                "序号",
                "巡视日期",
                "问题道路",
                "具体位置",
                "问题表述",
                "问题类型",
                "巡 视 问 题",
                "处 理 情 况",
                "备  注",
                "维修性质",
                "计划完成时间",
                "实际完成时间",
                "单位",
                "维修工程量",
                "截止时间",
            ]

            for c, v in enumerate(header, start=1):
                ws.cell(1, c).value = v

            ws.cell(2, 1).value = 1
            ws.cell(2, 3).value = "环湖西三路"
            ws.cell(2, 4).value = "近申港大道"
            ws.cell(2, 5).value = "绿化带有杂草"
            ws.cell(2, 15).value = "10"

            wb.save(excel_path)
            wb.close()

            xunshi_image = object()
            chuzhi_image = object()
            fake_image_map = {
                (1, 7): [xunshi_image],
                (2, 8): [chuzhi_image],
            }

            exported_names = []

            def fake_export(img, extracted_dir, row_num, col_num, index):
                name = "xunshi.jpg" if img is xunshi_image else "chuzhi.jpg"
                path = extracted_dir / name
                path.write_bytes(b"img")
                exported_names.append(name)
                return str(path)

            with patch.object(app_engine, "_build_image_anchor_map", return_value=fake_image_map), patch.object(
                app_engine, "_export_anchor_image", side_effect=fake_export
            ):
                result = build_session(excel_path)

            data = result.paths.data_tsv.read_text(encoding="utf-8")
            self.assertIn("xunshi.jpg", data)
            self.assertNotIn("chuzhi.jpg", data)
            self.assertEqual(exported_names, ["xunshi.jpg"])

            cleanup_session_dir(result.paths.session_dir)

    def test_patrol2_zip_drawing_strict_to_chuzhi_column(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "patrol_zip_drawing.xlsx"

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            header = [
                "序号",
                "巡视日期",
                "问题道路",
                "具体位置",
                "问题表述",
                "问题类型",
                "巡 视 问 题",
                "处 置 情 况",
                "备  注",
                "维修性质",
                "计划完成时间",
                "实际完成时间",
                "单位",
                "维修工程量",
                "整改描述",
                "截止时间",
            ]

            for c, v in enumerate(header, start=1):
                ws.cell(1, c).value = v

            ws.cell(2, 1).value = 1
            ws.cell(2, 3).value = "环湖西三路"
            ws.cell(2, 4).value = "近申港大道"
            ws.cell(2, 5).value = "绿化带有杂草"
            ws.cell(2, 15).value = "已除草"

            wb.save(excel_path)
            wb.close()

            fake_zip_map = {
                (2, 7): ["xl/media/xunshi.jpg"],
                (2, 8): ["xl/media/chuzhi.jpg"],
            }
            exported_media = []

            def fake_export(workbook_path, media_path, extracted_dir, row_num, col_num, index):
                exported_media.append((media_path, row_num, col_num))
                name = Path(media_path).name
                out = extracted_dir / name
                out.write_bytes(b"img")
                return str(out)

            with patch.object(app_engine, "_build_image_anchor_map", return_value={}), patch.object(
                app_engine, "_load_wps_cell_image_map", return_value={}
            ), patch.object(app_engine, "_load_zip_drawing_image_map", return_value=fake_zip_map), patch.object(
                app_engine, "_export_zip_media_image", side_effect=fake_export
            ):
                result = build_session(excel_path, fill_profile="patrol2")

            data = result.paths.data_tsv.read_text(encoding="utf-8")
            self.assertIn("chuzhi.jpg", data)
            self.assertNotIn("xunshi.jpg", data)
            self.assertEqual(exported_media, [("xl/media/chuzhi.jpg", 2, 8)])

            cleanup_session_dir(result.paths.session_dir)

    def test_patrol_wps_dispimg_strict_to_xunshi_column(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            excel_path = base / "patrol_wps.xlsx"

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            header = [
                "序号",
                "巡视日期",
                "问题道路",
                "具体位置",
                "问题表述",
                "问题类型",
                "巡 视 问 题",
                "处 理 情 况",
                "备  注",
                "维修性质",
                "计划完成时间",
                "实际完成时间",
                "单位",
                "维修工程量",
                "截止时间",
            ]

            for c, v in enumerate(header, start=1):
                ws.cell(1, c).value = v

            ws.cell(2, 1).value = 1
            ws.cell(2, 3).value = "环湖西三路"
            ws.cell(2, 4).value = "近申港大道"
            ws.cell(2, 5).value = "绿化带有杂草"
            ws.cell(2, 7).value = '=DISPIMG("ID_XUNSHI",1)'
            ws.cell(2, 8).value = '=DISPIMG("ID_CHUZHI",1)'
            ws.cell(2, 15).value = "10"

            wb.save(excel_path)
            wb.close()

            fake_map = {
                "ID_XUNSHI": "xl/media/image1.png",
                "ID_CHUZHI": "xl/media/image2.png",
            }
            exported_ids = []

            def fake_export(workbook_path, image_id, image_map, extracted_dir, row_num, col_num, index):
                exported_ids.append((image_id, row_num, col_num))
                name = "xunshi.png" if image_id == "ID_XUNSHI" else "chuzhi.png"
                path = extracted_dir / name
                path.write_bytes(b"img")
                return str(path)

            with patch.object(app_engine, "_build_image_anchor_map", return_value={}), patch.object(
                app_engine, "_load_wps_cell_image_map", return_value=fake_map
            ), patch.object(app_engine, "_export_wps_cell_image", side_effect=fake_export):
                result = build_session(excel_path)

            data = result.paths.data_tsv.read_text(encoding="utf-8")
            self.assertIn("xunshi.png", data)
            self.assertNotIn("chuzhi.png", data)
            self.assertEqual(exported_ids, [("ID_XUNSHI", 2, 7)])

            cleanup_session_dir(result.paths.session_dir)


if __name__ == "__main__":
    unittest.main()
