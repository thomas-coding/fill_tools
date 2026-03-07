from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


BASE_DIR = Path(__file__).resolve().parent
DATA_TSV_PATH = BASE_DIR / "wechat_form_data.tsv"
PROGRESS_TSV_PATH = BASE_DIR / "wechat_form_progress.tsv"
META_PATH = BASE_DIR / "wechat_source_meta.txt"
EXTRACT_DIR = BASE_DIR / "wechat_extracted_images"

SOURCE_CANDIDATES = ["1.excel", "1.xlsx", "wechat_form_test.xlsx"]


def norm(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value)
    return text.replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


def norm_header(value: object) -> str:
    return re.sub(r"\s+", "", norm(value))


def find_source_workbook() -> Path:
    for name in SOURCE_CANDIDATES:
        p = BASE_DIR / name
        if p.exists():
            return p

    xlsx_files = sorted(BASE_DIR.glob("*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]

    raise FileNotFoundError("当前目录未找到可用 Excel 文件")


def normalize_source_path(source_path: Path) -> Tuple[Path, str]:
    if source_path.suffix.lower() != ".excel":
        return source_path, ""

    xlsx_path = source_path.with_suffix(".xlsx")
    xlsx_path.write_bytes(source_path.read_bytes())
    note = f"检测到 {source_path.name}，已自动镜像为 {xlsx_path.name} 供脚本处理"
    return xlsx_path, note


def load_progress_map() -> Dict[int, Tuple[str, str]]:
    progress: Dict[int, Tuple[str, str]] = {}
    if not PROGRESS_TSV_PATH.exists():
        return progress

    lines = PROGRESS_TSV_PATH.read_text(encoding="utf-8").splitlines()
    for line in lines[1:]:
        if not line.strip():
            continue
        cols = line.split("\t")
        if len(cols) < 3:
            continue
        try:
            row_num = int(cols[0])
        except ValueError:
            continue
        progress[row_num] = (norm(cols[1]), norm(cols[2]))
    return progress


def read_status_from_simple(ws: Worksheet, row_num: int) -> Tuple[str, str]:
    return norm(ws.cell(row_num, 8).value), norm(ws.cell(row_num, 9).value)


def read_status_from_patrol(ws: Worksheet, row_num: int) -> Tuple[str, str]:
    return norm(ws.cell(row_num, 16).value), norm(ws.cell(row_num, 17).value)


def merge_status(excel_status: str, excel_time: str, old_status: str, old_time: str) -> Tuple[str, str]:
    # 以 Excel 为准：Excel 里清空状态即视为重置。
    if excel_status == "已填":
        merged_time = excel_time if excel_time else old_time
        return "已填", merged_time

    return "", ""


def row_has_content(ws: Worksheet, row_num: int) -> bool:
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(row_num, c).value):
            return True
    return False


def header_map_from_row(ws: Worksheet, row_num: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = norm_header(ws.cell(row_num, c).value)
        if key:
            mapping[key] = c
    return mapping


def build_image_anchor_map(ws: Worksheet) -> Dict[Tuple[int, int], List[object]]:
    image_map: Dict[Tuple[int, int], List[object]] = {}
    images = getattr(ws, "_images", [])
    for img in images:
        anchor = getattr(img, "anchor", None)
        from_obj = getattr(anchor, "_from", None)
        if from_obj is None:
            continue

        row_num = from_obj.row + 1
        col_num = from_obj.col + 1
        image_map.setdefault((row_num, col_num), []).append(img)
    return image_map


def export_anchor_image(img: object, row_num: int, col_num: int, index: int) -> str:
    EXTRACT_DIR.mkdir(parents=True, exist_ok=True)
    ext = norm(getattr(img, "format", "png")).lower() or "png"
    if ext == "jpeg":
        ext = "jpg"

    file_path = EXTRACT_DIR / f"r{row_num}_c{col_num}_{index}.{ext}"
    data_func = getattr(img, "_data", None)
    if callable(data_func):
        data = data_func()
        if isinstance(data, (bytes, bytearray)):
            file_path.write_bytes(bytes(data))
        else:
            return ""
    else:
        return ""
    return str(file_path)


def resolve_photo_path(raw_value: str) -> str:
    value = norm(raw_value)
    if not value:
        return ""

    p = Path(value)
    if p.exists():
        return str(p)

    p2 = BASE_DIR / value
    if p2.exists():
        return str(p2)

    return value


def find_fallback_extracted_image(row_num: int, col_num: int) -> str:
    if not EXTRACT_DIR.exists():
        return ""

    patterns = [
        f"r{row_num}_c{col_num}_*.png",
        f"r{row_num}_c{col_num}_*.jpg",
        f"r{row_num}_c{col_num}_*.jpeg",
        f"r{row_num}_c{col_num}_*.bmp",
    ]
    for pat in patterns:
        matches = sorted(EXTRACT_DIR.glob(pat))
        if matches:
            return str(matches[0])
    return ""


def parse_simple_sheet(ws: Worksheet, progress_map: Dict[int, Tuple[str, str]]) -> Tuple[List[str], List[str]]:
    data_lines: List[str] = []
    progress_lines: List[str] = []

    max_row = ws.max_row
    for row_num in range(2, max_row + 1):
        values = [norm(ws.cell(row_num, c).value) for c in range(1, 8)]
        if not any(values):
            continue

        excel_status, excel_time = read_status_from_simple(ws, row_num)
        old_status, old_time = progress_map.get(row_num, ("", ""))
        merged_status, merged_time = merge_status(excel_status, excel_time, old_status, old_time)

        data_lines.append(
            f"{row_num}\t{values[0]}\t{values[1]}\t{values[2]}\t{values[3]}\t{values[4]}\t{values[5]}\t{values[6]}\n"
        )
        progress_lines.append(f"{row_num}\t{merged_status}\t{merged_time}\n")

    return data_lines, progress_lines


def parse_patrol_sheet(ws: Worksheet, progress_map: Dict[int, Tuple[str, str]]) -> Tuple[List[str], List[str]]:
    data_lines: List[str] = []
    progress_lines: List[str] = []

    image_map = build_image_anchor_map(ws)

    row_num = 1
    while row_num <= ws.max_row:
        row_header = norm_header(ws.cell(row_num, 1).value)
        if row_header != "序号":
            row_num += 1
            continue

        header_map = header_map_from_row(ws, row_num)
        data_row = row_num + 1
        if data_row > ws.max_row or not row_has_content(ws, data_row):
            row_num += 1
            continue

        col_address = header_map.get("问题道路")
        col_section = header_map.get("具体位置")
        col_deadline = header_map.get("截止时间")
        col_desc = header_map.get("问题表述")
        col_photo = header_map.get("巡视问题")

        if not (col_address and col_section and col_deadline and col_desc and col_photo):
            row_num += 2
            continue

        address = norm(ws.cell(data_row, col_address).value)
        section = norm(ws.cell(data_row, col_section).value)
        deadline = norm(ws.cell(data_row, col_deadline).value)
        description = norm(ws.cell(data_row, col_desc).value)

        photo = resolve_photo_path(norm(ws.cell(data_row, col_photo).value))
        if photo and not Path(photo).exists():
            photo = ""
        if not photo:
            imgs = image_map.get((data_row, col_photo), [])
            if imgs:
                photo = export_anchor_image(imgs[0], data_row, col_photo, 1)
        if not photo:
            photo = find_fallback_extracted_image(data_row, col_photo)

        excel_status, excel_time = read_status_from_patrol(ws, data_row)
        old_status, old_time = progress_map.get(data_row, ("", ""))
        merged_status, merged_time = merge_status(excel_status, excel_time, old_status, old_time)

        data_lines.append(
            f"{data_row}\t{address}\t{section}\t{deadline}\t\t{description}\t{photo}\t\n"
        )
        progress_lines.append(f"{data_row}\t{merged_status}\t{merged_time}\n")

        row_num += 2

    return data_lines, progress_lines


def main() -> int:
    try:
        source_path = find_source_workbook()
    except FileNotFoundError as err:
        print(str(err))
        return 1

    source_path, source_note = normalize_source_path(source_path)

    progress_map = load_progress_map()

    wb = load_workbook(source_path)
    try:
        ws = wb.active
        assert ws is not None
        rows: List[str] = [
            "source_row\t问题地址\t问题路段\t截止时间(小时数)\t问题类别\t问题描述\t上传照片路径\t处置方式\n"
        ]
        progress_rows: List[str] = ["source_row\t状态\t提交时间\n"]

        first_row_headers = {norm_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
        patrol_mode = "问题道路" in first_row_headers and "具体位置" in first_row_headers and "问题表述" in first_row_headers

        if patrol_mode:
            data_lines, state_lines = parse_patrol_sheet(ws, progress_map)
            mode = "patrol"
        else:
            data_lines, state_lines = parse_simple_sheet(ws, progress_map)
            mode = "simple"

        rows.extend(data_lines)
        progress_rows.extend(state_lines)

        DATA_TSV_PATH.write_text("".join(rows), encoding="utf-8")
        PROGRESS_TSV_PATH.write_text("".join(progress_rows), encoding="utf-8")
        META_PATH.write_text(f"source={source_path}\nmode={mode}\n", encoding="utf-8")
    finally:
        wb.close()

    if len(rows) == 1:
        print("Excel 中没有可用数据行。")
        return 1

    print(f"来源文件: {source_path.name}")
    if source_note:
        print(source_note)
    print(f"已生成: {DATA_TSV_PATH}")
    print(f"记录数: {len(rows) - 1}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
