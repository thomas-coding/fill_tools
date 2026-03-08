from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app_engine import (
    FILL_PROFILE_PATROL1,
    FILL_PROFILE_PATROL2,
    build_session,
    cleanup_session_dir,
    norm,
    norm_header,
)


@dataclass
class DataRow:
    source_row: int
    address: str
    section: str
    deadline: str
    category: str
    description: str
    photo_path: str
    disposal: str


def read_data_rows(data_tsv: Path) -> List[DataRow]:
    rows: List[DataRow] = []
    if not data_tsv.exists():
        return rows

    for line in data_tsv.read_text(encoding="utf-8").splitlines()[1:]:
        if not line.strip():
            continue

        cols = line.split("\t")
        while len(cols) < 8:
            cols.append("")

        try:
            source_row = int(norm(cols[0]))
        except ValueError:
            continue

        rows.append(
            DataRow(
                source_row=source_row,
                address=norm(cols[1]),
                section=norm(cols[2]),
                deadline=norm(cols[3]),
                category=norm(cols[4]),
                description=norm(cols[5]),
                photo_path=norm(cols[6]),
                disposal=norm(cols[7]),
            )
        )
    return rows


def _row_has_content(ws: Worksheet, row_num: int) -> bool:
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(row_num, c).value):
            return True
    return False


def _header_map_from_row(ws: Worksheet, row_num: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = norm_header(ws.cell(row_num, c).value)
        if key:
            mapping[key] = c
    return mapping


def parse_patrol_source(source_excel: Path) -> Dict[int, Dict[str, str]]:
    expected: Dict[int, Dict[str, str]] = {}

    wb = load_workbook(source_excel, data_only=True)
    try:
        ws = wb.active
        assert ws is not None

        row_num = 1
        while row_num <= ws.max_row:
            if norm_header(ws.cell(row_num, 1).value) != "序号":
                row_num += 1
                continue

            data_row = row_num + 1
            if data_row > ws.max_row or not _row_has_content(ws, data_row):
                row_num += 1
                continue

            header_map = _header_map_from_row(ws, row_num)
            col_road = header_map.get("问题道路")
            col_location = header_map.get("具体位置")
            col_issue = header_map.get("问题表述")
            col_rectify = (
                header_map.get("整改描述")
                or header_map.get("整改情况")
                or header_map.get("处置描述")
            )

            expected[data_row] = {
                "road": norm(ws.cell(data_row, col_road).value) if col_road else "",
                "location": norm(ws.cell(data_row, col_location).value) if col_location else "",
                "issue": norm(ws.cell(data_row, col_issue).value) if col_issue else "",
                "rectify": norm(ws.cell(data_row, col_rectify).value) if col_rectify else "",
            }
            row_num += 2
    finally:
        wb.close()

    return expected


def validate_mapping(
    profile: str,
    rows: List[DataRow],
    expected: Dict[int, Dict[str, str]],
) -> List[str]:
    issues: List[str] = []

    def assert_equal(source_row: int, field_name: str, actual: str, wanted: str) -> None:
        if norm(actual) != norm(wanted):
            issues.append(
                f"行{source_row} 字段[{field_name}] 不匹配: 解析='{actual}' / 期望='{wanted}'"
            )

    for rec in rows:
        source = expected.get(rec.source_row)
        if source is None:
            issues.append(f"行{rec.source_row} 在源表中未找到对应记录")
            continue

        if profile == FILL_PROFILE_PATROL2:
            assert_equal(rec.source_row, "处置路段", rec.section, source.get("road", ""))
            rectify_or_issue = source.get("rectify", "") or source.get("issue", "")
            assert_equal(rec.source_row, "整改描述", rec.description, rectify_or_issue)
        else:
            assert_equal(rec.source_row, "问题地址", rec.address, source.get("road", ""))
            assert_equal(rec.source_row, "问题路段", rec.section, source.get("location", ""))
            assert_equal(rec.source_row, "问题描述", rec.description, source.get("issue", ""))

    return issues


def detect_image_type(path: Path) -> str:
    try:
        header = path.read_bytes()[:16]
    except Exception:
        return ""

    if len(header) >= 3 and header[:3] == b"\xff\xd8\xff":
        return "jpeg"
    if len(header) >= 8 and header[:8] == b"\x89PNG\r\n\x1a\n":
        return "png"
    if len(header) >= 6 and (header[:6] == b"GIF87a" or header[:6] == b"GIF89a"):
        return "gif"
    if len(header) >= 2 and header[:2] == b"BM":
        return "bmp"
    if len(header) >= 12 and header[:4] == b"RIFF" and header[8:12] == b"WEBP":
        return "webp"
    return ""


def validate_images(rows: List[DataRow]) -> Tuple[int, int, int]:
    ok = 0
    missing_or_empty = 0
    invalid = 0

    for rec in rows:
        photo = norm(rec.photo_path)
        if not photo:
            missing_or_empty += 1
            continue

        fp = Path(photo)
        if not fp.exists() or fp.stat().st_size <= 0:
            missing_or_empty += 1
            continue

        if detect_image_type(fp):
            ok += 1
        else:
            invalid += 1

    return ok, missing_or_empty, invalid


def profile_list(arg: str) -> List[str]:
    if arg == "all":
        return [FILL_PROFILE_PATROL1, FILL_PROFILE_PATROL2]
    if arg == FILL_PROFILE_PATROL2:
        return [FILL_PROFILE_PATROL2]
    return [FILL_PROFILE_PATROL1]


def run_profile(source_excel: Path, profile: str, sample_count: int, max_issues: int, keep_session: bool) -> bool:
    print(f"\n=== 离线自测 {profile} ===")
    result = build_session(source_excel, fill_profile=profile)
    rows = read_data_rows(result.paths.data_tsv)

    print(f"模式: {result.mode}")
    print(f"记录数: {result.total_records}")
    print(f"数据文件: {result.paths.data_tsv}")
    print(f"图片目录: {result.paths.extracted_dir}")

    show_count = max(0, min(sample_count, len(rows)))
    if show_count:
        print("样例记录:")
        for rec in rows[:show_count]:
            short_desc = rec.description if len(rec.description) <= 24 else rec.description[:24] + "..."
            print(
                f"- 行{rec.source_row}: 地址='{rec.address}' 路段='{rec.section}' 描述='{short_desc}'"
            )

    mapping_ok = True
    if result.mode == "patrol":
        expected = parse_patrol_source(source_excel)
        issues = validate_mapping(profile, rows, expected)
        if issues:
            mapping_ok = False
            print(f"字段校验: FAIL（{len(issues)}项）")
            for item in issues[:max_issues]:
                print(f"  * {item}")
            if len(issues) > max_issues:
                print(f"  * ... 其余 {len(issues) - max_issues} 项省略")
        else:
            print("字段校验: PASS")
    else:
        print("字段校验: 跳过（仅巡检模板支持此校验）")

    ok, missing_or_empty, invalid = validate_images(rows)
    image_ok = missing_or_empty == 0 and invalid == 0
    image_state = "PASS" if image_ok else "FAIL"
    print(
        f"图片校验: {image_state}（有效 {ok} / 缺失或空文件 {missing_or_empty} / 头信息异常 {invalid}）"
    )

    profile_ok = mapping_ok and image_ok
    print(f"结果: {'通过' if profile_ok else '失败'}")

    if not keep_session:
        cleanup_session_dir(result.paths.session_dir)
    else:
        print(f"已保留会话目录: {result.paths.session_dir}")

    return profile_ok


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="离线自测：校验巡检 Excel 的字段映射与图片提取，不依赖登录小程序。"
    )
    parser.add_argument("excel", type=Path, help="要测试的 Excel 路径")
    parser.add_argument(
        "--profile",
        choices=["all", FILL_PROFILE_PATROL1, FILL_PROFILE_PATROL2],
        default="all",
        help="测试功能类型，默认 all（填报1+填报2）",
    )
    parser.add_argument("--samples", type=int, default=3, help="输出样例记录条数，默认 3")
    parser.add_argument("--max-issues", type=int, default=8, help="最多展示多少条字段错误，默认 8")
    parser.add_argument("--keep-session", action="store_true", help="保留临时会话目录，便于人工查看提取图片")
    return parser.parse_args(argv)


def main(argv: List[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    source_excel = args.excel.resolve()

    if not source_excel.exists():
        print(f"文件不存在: {source_excel}")
        return 2

    ok = True
    for profile in profile_list(args.profile):
        try:
            profile_ok = run_profile(
                source_excel=source_excel,
                profile=profile,
                sample_count=args.samples,
                max_issues=args.max_issues,
                keep_session=args.keep_session,
            )
        except Exception as exc:
            ok = False
            print(f"\n=== 离线自测 {profile} ===")
            print(f"运行异常: {exc}")
            continue
        ok = ok and profile_ok

    print(f"\n总体结果: {'通过' if ok else '失败'}")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
