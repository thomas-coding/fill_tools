from __future__ import annotations

import json
import posixpath
import re
import shutil
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


OPENPYXL_WRITE_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUPPORTED_FILE_EXTS = {
    ".xlsx",
    ".xlsm",
    ".xltx",
    ".xltm",
    ".xls",
    ".xlsb",
    ".excel",
}

FILL_PROFILE_PATROL1 = "patrol1"
FILL_PROFILE_PATROL2 = "patrol2"
SUPPORTED_FILL_PROFILES = {FILL_PROFILE_PATROL1, FILL_PROFILE_PATROL2}


@dataclass
class Record:
    source_row: int
    address: str
    section: str
    deadline: str
    category: str
    description: str
    photo_path: str
    disposal: str
    done: bool = False
    submit_time: str = ""


@dataclass
class SessionPaths:
    session_dir: Path
    data_tsv: Path
    progress_tsv: Path
    meta_json: Path
    extracted_dir: Path
    ahk_script: Path


@dataclass
class SessionBuildResult:
    source_excel: Path
    mode: str
    fill_profile: str
    total_records: int
    paths: SessionPaths


def norm(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value)
    return text.replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


def norm_header(value: object) -> str:
    return re.sub(r"\s+", "", norm(value))


def is_supported_excel(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_FILE_EXTS


def normalize_fill_profile(fill_profile: str) -> str:
    profile = norm(fill_profile).lower() or FILL_PROFILE_PATROL1
    if profile not in SUPPORTED_FILL_PROFILES:
        raise ValueError(f"不支持的功能类型: {fill_profile}")
    return profile


def create_session_paths() -> SessionPaths:
    base = Path(tempfile.gettempdir()) / "patrol_form_assistant"
    base.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    session_dir = base / f"session_{stamp}"
    session_dir.mkdir(parents=True, exist_ok=True)

    extracted_dir = session_dir / "extracted_images"
    extracted_dir.mkdir(parents=True, exist_ok=True)

    return SessionPaths(
        session_dir=session_dir,
        data_tsv=session_dir / "wechat_form_data.tsv",
        progress_tsv=session_dir / "wechat_form_progress.tsv",
        meta_json=session_dir / "session_meta.json",
        extracted_dir=extracted_dir,
        ahk_script=session_dir / "runtime_helper.ahk",
    )


def cleanup_session_dir(session_dir: Path) -> None:
    try:
        shutil.rmtree(session_dir, ignore_errors=True)
    except Exception:
        pass


def _load_workbook_for_parse(source: Path, temp_dir: Path):
    try:
        wb = load_workbook(source)
        return wb, source, False
    except Exception:
        pass

    converted = _convert_with_excel_to_xlsx(source, temp_dir)
    wb = load_workbook(converted)
    return wb, converted, True


def _convert_with_excel_to_xlsx(source: Path, temp_dir: Path) -> Path:
    try:
        import win32com.client  # type: ignore
    except Exception as exc:  # pragma: no cover - requires Windows Excel env
        raise RuntimeError(
            "该文件格式需要本机安装 Excel 才能转换解析（支持 xls/xlsb/.excel）。"
        ) from exc

    converted = temp_dir / "converted_parse.xlsx"

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(source))
        wb.SaveAs(str(converted), FileFormat=51)
        wb.Close(False)
        wb = None
    finally:  # pragma: no cover - requires Windows Excel env
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        excel.Quit()

    return converted


def _detect_mode(ws: Worksheet) -> str:
    first_row_headers = {norm_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
    if "问题道路" in first_row_headers and "具体位置" in first_row_headers and "问题表述" in first_row_headers:
        return "patrol"
    return "simple"


def _build_image_anchor_map(ws: Worksheet) -> Dict[Tuple[int, int], List[object]]:
    image_map: Dict[Tuple[int, int], List[object]] = {}
    images = getattr(ws, "_images", [])
    for img in images:
        anchor = getattr(img, "anchor", None)
        from_obj = getattr(anchor, "_from", None)
        if from_obj is None:
            continue
        row_num = from_obj.row + 1
        col_num = from_obj.col + 1
        image_map.setdefault((row_num, col_num), []).append(img)
    return image_map


def _export_anchor_image(img: object, extracted_dir: Path, row_num: int, col_num: int, index: int) -> str:
    ext = norm(getattr(img, "format", "png")).lower() or "png"
    if ext == "jpeg":
        ext = "jpg"

    file_path = extracted_dir / f"r{row_num}_c{col_num}_{index}.{ext}"
    data_func = getattr(img, "_data", None)
    if not callable(data_func):
        return ""

    data = data_func()
    if not isinstance(data, (bytes, bytearray)):
        return ""

    file_path.write_bytes(bytes(data))
    return str(file_path)


def _extract_dispimg_id(value: object) -> str:
    text = norm(value)
    if not text:
        return ""

    m = re.search(r"DISPIMG\(\s*\"([^\"]+)\"", text, flags=re.IGNORECASE)
    if m:
        return norm(m.group(1))

    m = re.search(r"DISPIMG\(\s*'([^']+)'", text, flags=re.IGNORECASE)
    if m:
        return norm(m.group(1))

    return ""


def _load_wps_cell_image_map(workbook_path: Path) -> Dict[str, str]:
    cellimages_xml = "xl/cellimages.xml"
    cellimages_rels = "xl/_rels/cellimages.xml.rels"

    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    r_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    result: Dict[str, str] = {}
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            names = set(archive.namelist())
            if cellimages_xml not in names or cellimages_rels not in names:
                return {}

            rel_root = ET.fromstring(archive.read(cellimages_rels))
            rel_map: Dict[str, str] = {}
            for rel in rel_root.findall(f"{{{rel_ns}}}Relationship"):
                rel_id = norm(rel.attrib.get("Id"))
                target = norm(rel.attrib.get("Target"))
                if not rel_id or not target:
                    continue

                normalized = target.replace("\\", "/").lstrip("/")
                if not normalized.startswith("xl/"):
                    normalized = posixpath.join("xl", normalized)
                normalized = posixpath.normpath(normalized)
                rel_map[rel_id] = normalized

            root = ET.fromstring(archive.read(cellimages_xml))

            for cell_image in root.iter():
                if not str(cell_image.tag).endswith("}cellImage") and cell_image.tag != "cellImage":
                    continue

                c_nv_pr = None
                blip = None
                for elem in cell_image.iter():
                    tag = str(elem.tag)
                    if c_nv_pr is None and (tag.endswith("}cNvPr") or tag == "cNvPr"):
                        c_nv_pr = elem
                    if blip is None and (tag.endswith("}blip") or tag == "blip"):
                        blip = elem
                    if c_nv_pr is not None and blip is not None:
                        break

                if c_nv_pr is None or blip is None:
                    continue

                image_id = norm(c_nv_pr.attrib.get("name"))
                embed_id = ""
                for k, v in blip.attrib.items():
                    if str(k).endswith("}embed") or str(k) == "embed":
                        embed_id = norm(v)
                        break

                media_path = rel_map.get(embed_id, "")
                if image_id and media_path:
                    result[image_id] = media_path
    except Exception:
        return {}

    return result


def _xml_local_name(tag: object) -> str:
    text = str(tag)
    if "}" in text:
        return text.rsplit("}", 1)[-1]
    return text


def _load_zip_drawing_image_map(workbook_path: Path) -> Dict[Tuple[int, int], List[str]]:
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    result: Dict[Tuple[int, int], List[str]] = {}

    try:
        with zipfile.ZipFile(workbook_path) as archive:
            names = set(archive.namelist())
            drawing_files = sorted(
                n
                for n in names
                if n.startswith("xl/drawings/") and n.endswith(".xml") and "/_rels/" not in n
            )

            for drawing_xml in drawing_files:
                drawing_name = Path(drawing_xml).name
                rel_path = f"xl/drawings/_rels/{drawing_name}.rels"
                if rel_path not in names:
                    continue

                rel_root = ET.fromstring(archive.read(rel_path))
                rel_map: Dict[str, str] = {}
                for rel in rel_root.findall(f"{{{rel_ns}}}Relationship"):
                    rel_id = norm(rel.attrib.get("Id"))
                    target = norm(rel.attrib.get("Target"))
                    if not rel_id or not target:
                        continue

                    target_norm = target.replace("\\", "/")
                    if target_norm.startswith("/"):
                        normalized = target_norm.lstrip("/")
                    else:
                        normalized = posixpath.normpath(posixpath.join(posixpath.dirname(drawing_xml), target_norm))
                    rel_map[rel_id] = normalized

                root = ET.fromstring(archive.read(drawing_xml))
                for anchor in list(root):
                    anchor_name = _xml_local_name(anchor.tag)
                    if anchor_name not in {"oneCellAnchor", "twoCellAnchor", "absoluteAnchor"}:
                        continue

                    row_col: Tuple[int, int] | None = None
                    embed_id = ""
                    for elem in anchor.iter():
                        name = _xml_local_name(elem.tag)

                        if row_col is None and name == "from":
                            row_txt = ""
                            col_txt = ""
                            for child in list(elem):
                                child_name = _xml_local_name(child.tag)
                                if child_name == "row":
                                    row_txt = norm(child.text)
                                elif child_name == "col":
                                    col_txt = norm(child.text)

                            if row_txt.isdigit() and col_txt.isdigit():
                                row_col = (int(row_txt) + 1, int(col_txt) + 1)

                        if not embed_id and name == "blip":
                            for k, v in elem.attrib.items():
                                if str(k).endswith("}embed") or str(k) == "embed":
                                    embed_id = norm(v)
                                    break

                        if row_col and embed_id:
                            break

                    if not row_col or not embed_id:
                        continue

                    media_path = rel_map.get(embed_id, "")
                    if not media_path:
                        continue

                    result.setdefault(row_col, []).append(media_path)
    except Exception:
        return {}

    return result


def _export_zip_media_image(
    workbook_path: Path,
    media_path: str,
    extracted_dir: Path,
    row_num: int,
    col_num: int,
    index: int,
) -> str:
    media_file = norm(media_path)
    if not media_file:
        return ""

    ext = Path(media_file).suffix.lower().lstrip(".") or "png"
    if ext == "jpeg":
        ext = "jpg"

    file_path = extracted_dir / f"r{row_num}_c{col_num}_{index}.{ext}"
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            data = archive.read(media_file)
    except Exception:
        return ""

    if not data:
        return ""

    file_path.write_bytes(data)
    return str(file_path)


def _export_wps_cell_image(
    workbook_path: Path,
    image_id: str,
    image_map: Dict[str, str],
    extracted_dir: Path,
    row_num: int,
    col_num: int,
    index: int,
) -> str:
    media_path = image_map.get(norm(image_id), "")
    if not media_path:
        return ""

    ext = Path(media_path).suffix.lower().lstrip(".") or "png"
    if ext == "jpeg":
        ext = "jpg"

    file_path = extracted_dir / f"r{row_num}_c{col_num}_{index}.{ext}"
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            data = archive.read(media_path)
    except Exception:
        return ""

    if not data:
        return ""

    file_path.write_bytes(data)
    return str(file_path)


def _resolve_photo_path(raw_value: str, source_dir: Path) -> str:
    value = norm(raw_value)
    if not value:
        return ""

    p = Path(value)
    if p.exists():
        return str(p)

    p2 = source_dir / value
    if p2.exists():
        return str(p2)

    return value


def _find_fallback_extracted_image(source_dir: Path, row_num: int, col_num: int) -> str:
    extracted_dir = source_dir / "wechat_extracted_images"
    if not extracted_dir.exists():
        return ""

    patterns = [
        f"r{row_num}_c{col_num}_*.png",
        f"r{row_num}_c{col_num}_*.jpg",
        f"r{row_num}_c{col_num}_*.jpeg",
        f"r{row_num}_c{col_num}_*.bmp",
    ]
    for pat in patterns:
        matches = sorted(extracted_dir.glob(pat))
        if matches:
            return str(matches[0])
    return ""


def _header_map_from_row(ws: Worksheet, row_num: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = norm_header(ws.cell(row_num, c).value)
        if key:
            mapping[key] = c
    return mapping


def _row_has_content(ws: Worksheet, row_num: int) -> bool:
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(row_num, c).value):
            return True
    return False


def _patrol_status_headers(fill_profile: str) -> Tuple[str, str]:
    if fill_profile == FILL_PROFILE_PATROL2:
        return "填报2状态", "填报2提交时间"
    return "填报状态", "提交时间"


def _patrol_status_cols_from_header_map(header_map: Dict[str, int], fill_profile: str) -> Tuple[int, int]:
    deadline_col = header_map.get("截止时间") or 0

    if fill_profile == FILL_PROFILE_PATROL2:
        if deadline_col:
            return deadline_col + 3, deadline_col + 4

        status_col = header_map.get("填报2状态") or header_map.get("巡检填报2状态")
        submit_col = header_map.get("填报2提交时间") or header_map.get("巡检填报2提交时间")
        if not submit_col and status_col:
            submit_col = status_col + 1
        return status_col or 0, submit_col or 0

    if deadline_col:
        return deadline_col + 1, deadline_col + 2

    status_col = header_map.get("填报状态") or header_map.get("状态")
    submit_col = header_map.get("提交时间")
    if not status_col:
        status_col = 16
    if not submit_col:
        submit_col = status_col + 1

    return status_col, submit_col


def _patrol_status_cols_openpyxl(
    ws: Worksheet,
    row_num: int,
    fill_profile: str,
    create_if_missing: bool = False,
) -> Tuple[int, int]:
    header_row = max(1, row_num - 1)
    header_map = _header_map_from_row(ws, header_row)
    status_col, submit_col = _patrol_status_cols_from_header_map(header_map, fill_profile)

    if fill_profile == FILL_PROFILE_PATROL2 and create_if_missing and (status_col <= 0 or submit_col <= 0):
        status_header, submit_header = _patrol_status_headers(fill_profile)
        next_col = max(header_map.values(), default=0)
        if next_col <= 0:
            next_col = ws.max_column
        if status_col <= 0:
            next_col += 1
            status_col = next_col
            _safe_set(ws, header_row, status_col, status_header)
        if submit_col <= 0:
            next_col = max(next_col, status_col)
            next_col += 1
            submit_col = next_col
            _safe_set(ws, header_row, submit_col, submit_header)

    if fill_profile == FILL_PROFILE_PATROL2 and (status_col <= 0 or submit_col <= 0):
        return 0, 0

    return status_col, submit_col


def _status_from_excel(ws: Worksheet, row_num: int, mode: str, fill_profile: str) -> Tuple[str, str]:
    if mode == "patrol":
        status_col, submit_col = _patrol_status_cols_openpyxl(ws, row_num, fill_profile, create_if_missing=False)
        if status_col <= 0 or submit_col <= 0:
            return "", ""
        return norm(ws.cell(row_num, status_col).value), norm(ws.cell(row_num, submit_col).value)
    return norm(ws.cell(row_num, 8).value), norm(ws.cell(row_num, 9).value)


def _parse_simple(ws: Worksheet, source_dir: Path) -> List[Record]:
    records: List[Record] = []
    for row_num in range(2, ws.max_row + 1):
        values = [norm(ws.cell(row_num, c).value) for c in range(1, 8)]
        if not any(values):
            continue

        records.append(
            Record(
                source_row=row_num,
                address=values[0],
                section=values[1],
                deadline=values[2],
                category=values[3],
                description=values[4],
                photo_path=_resolve_photo_path(values[5], source_dir),
                disposal=values[6],
            )
        )
    return records


def _parse_patrol(
    ws: Worksheet,
    source_dir: Path,
    extracted_dir: Path,
    workbook_path: Path,
    fill_profile: str,
) -> List[Record]:
    records: List[Record] = []
    image_map = _build_image_anchor_map(ws)
    wps_image_map = _load_wps_cell_image_map(workbook_path)
    zip_drawing_map = _load_zip_drawing_image_map(workbook_path)
    fill_profile = normalize_fill_profile(fill_profile)

    row_num = 1
    while row_num <= ws.max_row:
        if norm_header(ws.cell(row_num, 1).value) != "序号":
            row_num += 1
            continue

        data_row = row_num + 1
        if data_row > ws.max_row or not _row_has_content(ws, data_row):
            row_num += 1
            continue

        header_map = _header_map_from_row(ws, row_num)
        col_road = header_map.get("问题道路")
        col_location = header_map.get("具体位置")
        col_deadline = header_map.get("截止时间")
        col_issue_desc = header_map.get("问题表述")
        col_rectify_desc = (
            header_map.get("整改描述")
            or header_map.get("整改情况")
            or header_map.get("处置描述")
        )
        col_patrol_photo = header_map.get("巡视问题")
        col_disposal_photo = header_map.get("处置情况") or header_map.get("处理情况")

        if fill_profile == FILL_PROFILE_PATROL2:
            col_section = col_road
            col_desc = col_rectify_desc or col_issue_desc
            col_photo = col_disposal_photo
        else:
            col_section = col_location
            col_desc = col_issue_desc
            col_photo = col_patrol_photo

        if not (col_road and col_section and col_desc and col_photo):
            row_num += 2
            continue

        photo = _resolve_photo_path(norm(ws.cell(data_row, col_photo).value), source_dir)
        if photo and not Path(photo).exists():
            photo = ""

        if not photo:
            imgs = image_map.get((data_row, col_photo), [])
            if not imgs:
                # Keep matching strict to the selected source photo column only.
                # Allow a tiny row drift (saved by different Excel clients),
                # but never fall back to other columns.
                for fallback_row in (data_row - 1, data_row + 1):
                    if fallback_row <= 0:
                        continue
                    imgs = image_map.get((fallback_row, col_photo), [])
                    if imgs:
                        break

            if imgs:
                photo = _export_anchor_image(imgs[0], extracted_dir, data_row, col_photo, 1)

        if not photo and wps_image_map:
            dispimg_id = _extract_dispimg_id(ws.cell(data_row, col_photo).value)
            if dispimg_id:
                photo = _export_wps_cell_image(
                    workbook_path,
                    dispimg_id,
                    wps_image_map,
                    extracted_dir,
                    data_row,
                    col_photo,
                    1,
                )

        if not photo and zip_drawing_map:
            media_paths = zip_drawing_map.get((data_row, col_photo), [])
            if not media_paths:
                for fallback_row in (data_row - 1, data_row + 1):
                    if fallback_row <= 0:
                        continue
                    media_paths = zip_drawing_map.get((fallback_row, col_photo), [])
                    if media_paths:
                        break

            if media_paths:
                photo = _export_zip_media_image(
                    workbook_path,
                    media_paths[0],
                    extracted_dir,
                    data_row,
                    col_photo,
                    1,
                )

        if not photo:
            photo = _find_fallback_extracted_image(source_dir, data_row, col_photo)

        address_value = ""
        if col_road is not None:
            address_value = norm(ws.cell(data_row, col_road).value)
        elif col_location is not None:
            address_value = norm(ws.cell(data_row, col_location).value)

        records.append(
            Record(
                source_row=data_row,
                address=address_value,
                section=norm(ws.cell(data_row, col_section).value),
                deadline=norm(ws.cell(data_row, col_deadline).value)
                if (col_deadline and fill_profile != FILL_PROFILE_PATROL2)
                else "",
                category="",
                description=norm(ws.cell(data_row, col_desc).value),
                photo_path=photo,
                disposal="",
            )
        )

        row_num += 2

    return records


def build_session(source_excel: Path, fill_profile: str = FILL_PROFILE_PATROL1) -> SessionBuildResult:
    fill_profile = normalize_fill_profile(fill_profile)
    source_excel = source_excel.resolve()
    if not source_excel.exists():
        raise FileNotFoundError(f"文件不存在: {source_excel}")
    if not is_supported_excel(source_excel):
        raise ValueError("不支持的文件格式，请选择 xlsx/xlsm/xls/xlsb/.excel 文件")

    paths = create_session_paths()
    try:
        wb, parsed_path, used_conversion = _load_workbook_for_parse(source_excel, paths.session_dir)
        try:
            ws = wb.active
            assert ws is not None

            mode = _detect_mode(ws)
            if mode == "patrol":
                records = _parse_patrol(ws, source_excel.parent, paths.extracted_dir, parsed_path, fill_profile)
            else:
                records = _parse_simple(ws, source_excel.parent)

            for rec in records:
                status, submit_time = _status_from_excel(ws, rec.source_row, mode, fill_profile)
                rec.done = status == "已填"
                rec.submit_time = submit_time if rec.done else ""
        finally:
            wb.close()

        if not records:
            raise RuntimeError("未解析到可用记录，请检查 Excel 模板是否符合要求")

        write_data_tsv(paths.data_tsv, records)
        write_progress_tsv(paths.progress_tsv, records)

        meta = {
            "source_excel": str(source_excel),
            "mode": mode,
            "fill_profile": fill_profile,
            "used_conversion_for_parse": used_conversion,
            "parsed_workbook": str(parsed_path),
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        paths.meta_json.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

        return SessionBuildResult(
            source_excel=source_excel,
            mode=mode,
            fill_profile=fill_profile,
            total_records=len(records),
            paths=paths,
        )
    except Exception:
        cleanup_session_dir(paths.session_dir)
        raise


def write_data_tsv(path: Path, records: Iterable[Record]) -> None:
    lines = [
        "source_row\t问题地址\t问题路段\t截止时间(小时数)\t问题类别\t问题描述\t上传照片路径\t处置方式\n"
    ]
    for rec in records:
        lines.append(
            f"{rec.source_row}\t{rec.address}\t{rec.section}\t{rec.deadline}\t{rec.category}\t"
            f"{rec.description}\t{rec.photo_path}\t{rec.disposal}\n"
        )
    path.write_text("".join(lines), encoding="utf-8")


def write_progress_tsv(path: Path, records: Iterable[Record]) -> None:
    lines = ["source_row\t状态\t提交时间\n"]
    for rec in records:
        status = "已填" if rec.done else ""
        submit_time = rec.submit_time if rec.done else ""
        lines.append(f"{rec.source_row}\t{status}\t{submit_time}\n")
    path.write_text("".join(lines), encoding="utf-8")


def read_progress_tsv(path: Path) -> Dict[int, Tuple[str, str]]:
    progress: Dict[int, Tuple[str, str]] = {}
    if not path.exists():
        return progress

    lines = path.read_text(encoding="utf-8").splitlines()
    for line in lines[1:]:
        if not line.strip():
            continue
        cols = line.split("\t")
        while len(cols) < 3:
            cols.append("")
        try:
            row_num = int(cols[0])
        except ValueError:
            continue
        progress[row_num] = (norm(cols[1]), norm(cols[2]))
    return progress


def progress_stats(path: Path) -> Tuple[int, int]:
    progress = read_progress_tsv(path)
    total = len(progress)
    done = sum(1 for status, _ in progress.values() if status == "已填")
    return done, total


def _safe_set(ws: Worksheet, row_num: int, col_num: int, value: str) -> None:
    try:
        cell = ws.cell(row_num, col_num)
        setattr(cell, "value", value)
    except Exception:
        pass


def _header_map_from_com_row(ws: Any, row_num: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    try:
        max_col = int(ws.UsedRange.Columns.Count)
    except Exception:
        max_col = 80

    max_col = max(max_col, 20)
    for c in range(1, max_col + 1):
        try:
            key = norm_header(ws.Cells(row_num, c).Value)
        except Exception:
            key = ""
        if key:
            mapping[key] = c
    return mapping


def _patrol_status_cols_com(ws: Any, row_num: int, fill_profile: str, create_if_missing: bool) -> Tuple[int, int]:
    header_row = max(1, row_num - 1)
    header_map = _header_map_from_com_row(ws, header_row)
    status_col, submit_col = _patrol_status_cols_from_header_map(header_map, fill_profile)

    if fill_profile == FILL_PROFILE_PATROL2 and create_if_missing and (status_col <= 0 or submit_col <= 0):
        status_header, submit_header = _patrol_status_headers(fill_profile)
        next_col = max(max(header_map.values(), default=0), status_col, submit_col)
        if next_col <= 0:
            try:
                next_col = int(ws.UsedRange.Columns.Count)
            except Exception:
                next_col = 0
        if status_col <= 0:
            next_col += 1
            status_col = next_col
            ws.Cells(header_row, status_col).Value = status_header
        if submit_col <= 0:
            next_col = max(next_col, status_col)
            next_col += 1
            submit_col = next_col
            ws.Cells(header_row, submit_col).Value = submit_header

    if fill_profile == FILL_PROFILE_PATROL2 and (status_col <= 0 or submit_col <= 0):
        return 0, 0

    return status_col, submit_col


def sync_progress_to_source(meta_json: Path, progress_tsv: Path) -> int:
    if not meta_json.exists():
        return 0

    meta = json.loads(meta_json.read_text(encoding="utf-8"))
    source_excel = Path(meta.get("source_excel", ""))
    mode = str(meta.get("mode", "simple"))
    fill_profile = normalize_fill_profile(str(meta.get("fill_profile", FILL_PROFILE_PATROL1)))
    if not source_excel.exists():
        return 0

    progress_map = read_progress_tsv(progress_tsv)
    if not progress_map:
        return 0

    if mode == "patrol":
        # 巡检模板通常含嵌图，优先COM回写，避免openpyxl回写导致嵌图丢失。
        try:
            return _sync_by_excel_com(source_excel, mode, progress_map, fill_profile)
        except Exception:
            if _workbook_has_embedded_images(source_excel):
                # 若有嵌图且COM不可用，跳过回写以保护源文件。
                return 0
            # 无嵌图时可降级为 openpyxl。
            return _sync_by_openpyxl(source_excel, mode, progress_map, fill_profile)

    try:
        updated = _sync_by_openpyxl(source_excel, mode, progress_map, fill_profile)
        return updated
    except Exception:
        return _sync_by_excel_com(source_excel, mode, progress_map, fill_profile)


def _sync_by_openpyxl(
    source_excel: Path,
    mode: str,
    progress_map: Dict[int, Tuple[str, str]],
    fill_profile: str,
) -> int:
    if source_excel.suffix.lower() not in OPENPYXL_WRITE_EXTS:
        raise RuntimeError("source requires COM sync")

    wb = load_workbook(source_excel)
    updated = 0
    try:
        ws = wb.active
        assert ws is not None
        if mode == "patrol":
            status_header, submit_header = _patrol_status_headers(fill_profile)
            for row_num, (status, submit_time) in progress_map.items():
                if row_num <= 1:
                    continue
                status_col, submit_col = _patrol_status_cols_openpyxl(
                    ws,
                    row_num,
                    fill_profile,
                    create_if_missing=(fill_profile == FILL_PROFILE_PATROL2),
                )
                if status_col <= 0 or submit_col <= 0:
                    continue

                _safe_set(ws, row_num - 1, status_col, status_header)
                _safe_set(ws, row_num - 1, submit_col, submit_header)
                _safe_set(ws, row_num, status_col, status)
                _safe_set(ws, row_num, submit_col, submit_time)
                updated += 1
        else:
            if not norm(ws.cell(1, 8).value):
                _safe_set(ws, 1, 8, "状态")
            if not norm(ws.cell(1, 9).value):
                _safe_set(ws, 1, 9, "提交时间")
            for row_num, (status, submit_time) in progress_map.items():
                _safe_set(ws, row_num, 8, status)
                _safe_set(ws, row_num, 9, submit_time)
                updated += 1

        wb.save(source_excel)
    finally:
        wb.close()

    return updated


def _workbook_has_embedded_images(source_excel: Path) -> bool:
    try:
        wb = load_workbook(source_excel)
    except Exception:
        return False

    try:
        ws = wb.active
        if ws is None:
            return False
        images = getattr(ws, "_images", [])
        return len(images) > 0
    finally:
        wb.close()


def _sync_by_excel_com(
    source_excel: Path,
    mode: str,
    progress_map: Dict[int, Tuple[str, str]],
    fill_profile: str,
) -> int:
    try:
        import win32com.client  # type: ignore
    except Exception as exc:  # pragma: no cover - requires Windows Excel env
        raise RuntimeError("当前文件格式需要 Excel COM 回写，但环境不可用") from exc

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    updated = 0
    try:
        wb = excel.Workbooks.Open(str(source_excel))
        ws = wb.Worksheets(1)

        if mode == "patrol":
            status_header, submit_header = _patrol_status_headers(fill_profile)
            for row_num, (status, submit_time) in progress_map.items():
                if row_num <= 1:
                    continue
                status_col, submit_col = _patrol_status_cols_com(
                    ws,
                    row_num,
                    fill_profile,
                    create_if_missing=(fill_profile == FILL_PROFILE_PATROL2),
                )
                if status_col <= 0 or submit_col <= 0:
                    continue

                ws.Cells(row_num - 1, status_col).Value = status_header
                ws.Cells(row_num - 1, submit_col).Value = submit_header
                ws.Cells(row_num, status_col).Value = status
                ws.Cells(row_num, submit_col).Value = submit_time
                updated += 1
        else:
            if not norm(ws.Cells(1, 8).Value):
                ws.Cells(1, 8).Value = "状态"
            if not norm(ws.Cells(1, 9).Value):
                ws.Cells(1, 9).Value = "提交时间"
            for row_num, (status, submit_time) in progress_map.items():
                ws.Cells(row_num, 8).Value = status
                ws.Cells(row_num, 9).Value = submit_time
                updated += 1

        wb.Save()
        wb.Close(False)
        wb = None
    finally:  # pragma: no cover - requires Windows Excel env
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        excel.Quit()

    return updated


def write_runtime_ahk(paths: SessionPaths, fill_profile: str = FILL_PROFILE_PATROL1) -> Path:
    script = _render_ahk_script(paths.data_tsv, paths.progress_tsv, fill_profile)
    paths.ahk_script.write_text(script, encoding="utf-8")
    return paths.ahk_script


def _ahk_path_literal(path: Path) -> str:
    # AHK v2 supports standard windows path with single backslashes in quoted string.
    return str(path)


def _render_ahk_script(data_tsv: Path, progress_tsv: Path, fill_profile: str) -> str:
    data_literal = _ahk_path_literal(data_tsv)
    progress_literal = _ahk_path_literal(progress_tsv)
    profile = normalize_fill_profile(fill_profile)
    default_skip_tabs = 0

    return f'''#Requires AutoHotkey v2.0
#SingleInstance Force
Persistent
SetWorkingDir A_ScriptDir

global CONFIG := {{
    dataFile: "{data_literal}",
    progressFile: "{progress_literal}",
    tabDelayMs: 90,
    pasteDelayMs: 70,
    skipTabsBeforeDescription: {default_skip_tabs},
    formProfile: "{profile}"
}}

global records := []
global currentIndex := 1

Init()

F5::FillTextFieldsOnce()
F8::PasteDescriptionOnly()
F12::FillPhotoInFileDialog()
F10::MarkCurrentDoneAndNext()
F9::GoNextPendingRow()
F6::GoPrevPendingRow()
F3::IncreaseSkipTabs()
F2::DecreaseSkipTabs()
F4::ShowCurrentRowValues()
F1::SaveAndExit()
^Esc::ExitApp

Init() {{
    global records, CONFIG

    if !FileExist(CONFIG.dataFile) {{
        MsgBox "未找到数据文件:`n" CONFIG.dataFile
        ExitApp
    }}

    records := LoadRecords()
    if records.Length = 0 {{
        MsgBox "数据文件为空，请检查 wechat_form_data.tsv。"
        ExitApp
    }}

    JumpToFirstPending()
    ShowHelp()
    OnExit(SaveProgress)
}}

ShowHelp() {{
    helpText := "微信小程序填报助手热键：`n"
    helpText .= "F5  巡检填报1填4项，巡检填报2填2项（处置路段/整改描述）`n"
    helpText .= "F8  仅粘贴问题描述（手动点到描述框后用）`n"
    helpText .= "F12 在文件选择框自动填入照片路径`n"
    helpText .= "F10 标记当前记录已填并跳下一条`n"
    helpText .= "F9  跳到下一条待填记录`n"
    helpText .= "F6  回到上一条待填记录`n"
    helpText .= "F3  +1 描述前Tab跳过数（当前: " CONFIG.skipTabsBeforeDescription "）`n"
    helpText .= "F2  -1 描述前Tab跳过数`n"
    helpText .= "F4  查看当前记录全部字段`n"
    helpText .= "F1  保存进度并退出`n"
    helpText .= "Ctrl+Esc 退出助手"
    MsgBox helpText
}}

LoadRecords() {{
    global CONFIG

    progressMap := LoadProgressMap()
    arr := []

    text := FileRead(CONFIG.dataFile, "UTF-8")
    lines := StrSplit(text, "`n", "`r")
    if lines.Length <= 1 {{
        return arr
    }}

    Loop lines.Length - 1 {{
        raw := Trim(lines[A_Index + 1], "`r`n")
        if raw = "" {{
            continue
        }}

        cols := StrSplit(raw, "`t")
        while cols.Length < 8 {{
            cols.Push("")
        }}

        srcRow := SafeInt(cols[1], A_Index + 1)
        rec := {{
            sourceRow: srcRow,
            address: cols[2],
            section: cols[3],
            deadlineHours: cols[4],
            category: cols[5],
            description: cols[6],
            photoPath: cols[7],
            disposal: cols[8],
            done: false,
            submitTime: ""
        }}

        if progressMap.Has(srcRow) {{
            p := progressMap[srcRow]
            rec.done := (p.status = "已填")
            rec.submitTime := p.submitTime
        }}

        arr.Push(rec)
    }}

    return arr
}}

LoadProgressMap() {{
    global CONFIG

    progressMap := Map()
    if !FileExist(CONFIG.progressFile) {{
        return progressMap
    }}

    text := FileRead(CONFIG.progressFile, "UTF-8")
    lines := StrSplit(text, "`n", "`r")
    if lines.Length <= 1 {{
        return progressMap
    }}

    Loop lines.Length - 1 {{
        raw := Trim(lines[A_Index + 1], "`r`n")
        if raw = "" {{
            continue
        }}

        cols := StrSplit(raw, "`t")
        while cols.Length < 3 {{
            cols.Push("")
        }}

        rowNum := SafeInt(cols[1], 0)
        if rowNum <= 0 {{
            continue
        }}

        progressMap[rowNum] := {{status: cols[2], submitTime: cols[3]}}
    }}

    return progressMap
}}

SaveProgress(*) {{
    global records, CONFIG

    text := "source_row`t状态`t提交时间`n"
    for rec in records {{
        status := rec.done ? "已填" : ""
        text .= rec.sourceRow "`t" status "`t" rec.submitTime "`n"
    }}

    try FileDelete(CONFIG.progressFile)
    FileAppend(text, CONFIG.progressFile, "UTF-8")
}}

JumpToFirstPending() {{
    global currentIndex, records

    currentIndex := 1
    for idx, rec in records {{
        if !rec.done {{
            currentIndex := idx
            Toast("当前源行: " rec.sourceRow "（待填）")
            return
        }}
    }}

    Toast("全部记录已填，按 F6 可回看")
}}

MoveToPending(direction := 1) {{
    global currentIndex, records

    idx := currentIndex + direction
    while idx >= 1 && idx <= records.Length {{
        if !records[idx].done {{
            currentIndex := idx
            Toast("当前源行: " records[idx].sourceRow "（待填）")
            return true
        }}
        idx += direction
    }}

    return false
}}

GoNextPendingRow() {{
    if !MoveToPending(1) {{
        Toast("后面没有待填记录")
    }}
}}

GoPrevPendingRow() {{
    if !MoveToPending(-1) {{
        Toast("前面没有待填记录")
    }}
}}

FillTextFieldsOnce() {{
    global records, currentIndex, CONFIG

    rec := records[currentIndex]

    if CONFIG.formProfile = "{FILL_PROFILE_PATROL2}" {{
        if !PasteValue(rec.section) {{
            return
        }}
        Send "{{Tab}}"
        Sleep CONFIG.tabDelayMs

        if CONFIG.skipTabsBeforeDescription > 0 {{
            Send "{{Tab " CONFIG.skipTabsBeforeDescription "}}"
            Sleep CONFIG.tabDelayMs
        }}

        if !PasteValue(rec.description) {{
            return
        }}

        Toast("源行 " rec.sourceRow " 已填2项（处置路段/整改描述，跳过Tab=" CONFIG.skipTabsBeforeDescription "）")
        return
    }}

    if !PasteValue(rec.address) {{
        return
    }}
    Send "{{Tab}}"
    Sleep CONFIG.tabDelayMs

    if !PasteValue(rec.section) {{
        return
    }}
    Send "{{Tab}}"
    Sleep CONFIG.tabDelayMs

    if !PasteValue(rec.deadlineHours) {{
        return
    }}

    Send "{{Tab}}"
    Sleep CONFIG.tabDelayMs
    if CONFIG.skipTabsBeforeDescription > 0 {{
        Send "{{Tab " CONFIG.skipTabsBeforeDescription "}}"
        Sleep CONFIG.tabDelayMs
    }}

    if !PasteValue(rec.description) {{
        return
    }}

    Toast("源行 " rec.sourceRow " 四项文本已填（跳过Tab=" CONFIG.skipTabsBeforeDescription "）")
}}

PasteDescriptionOnly() {{
    global records, currentIndex

    rec := records[currentIndex]
    if !PasteValue(rec.description) {{
        return
    }}
    Toast("源行 " rec.sourceRow " 问题描述已粘贴")
}}

IncreaseSkipTabs() {{
    global CONFIG
    CONFIG.skipTabsBeforeDescription += 1
    Toast("描述前Tab跳过数: " CONFIG.skipTabsBeforeDescription)
}}

DecreaseSkipTabs() {{
    global CONFIG
    if CONFIG.skipTabsBeforeDescription > 0 {{
        CONFIG.skipTabsBeforeDescription -= 1
    }}
    Toast("描述前Tab跳过数: " CONFIG.skipTabsBeforeDescription)
}}

PasteValue(value) {{
    global CONFIG

    A_Clipboard := value
    if !ClipWait(0.5) {{
        Toast("复制到剪贴板失败")
        return false
    }}

    Send "^v"
    Sleep CONFIG.pasteDelayMs
    return true
}}

FillPhotoInFileDialog() {{
    global records, currentIndex

    rawPath := records[currentIndex].photoPath
    fullPath := ResolvePhotoPath(rawPath)
    if fullPath = "" {{
        Toast("照片不存在或路径为空: " rawPath)
        return
    }}

    if !WaitForAnyFileDialog(2500) {{
        MsgBox "未检测到文件选择框。`n请先点击上传照片 +，确保文件框处于前台后再按 F12。"
        return
    }}

    if TryFillFileDialogByControl(fullPath) {{
        Toast("已按数据路径提交文件")
        return
    }}

    A_Clipboard := fullPath
    if !ClipWait(0.5) {{
        Toast("复制照片路径失败")
        return
    }}

    Send "!n"
    Sleep 80
    Send "^a^v{{Enter}}"
    Sleep 220

    if IsAnyFileDialogActive() {{
        MsgBox "已尝试自动输入路径，但文件框仍未关闭。`n请检查图片路径是否有效: `n" fullPath
        return
    }}

    Toast("已按数据路径提交文件")
}}

TryFillFileDialogByControl(fullPath) {{
    try {{
        ControlFocus "Edit1", "A"
        Sleep 60
        ControlSetText fullPath, "Edit1", "A"
        Sleep 80
        ControlSend "{{Enter}}", "Edit1", "A"
        Sleep 220
        return !IsAnyFileDialogActive()
    }} catch {{
        return false
    }}
}}

IsAnyFileDialogActive() {{
    return WinActive("ahk_class #32770")
        || WinActive("ahk_class CabinetWClass")
        || WinActive("ahk_class ExploreWClass")
        || WinActive("打开")
        || WinActive("Open")
}}

WaitForAnyFileDialog(timeoutMs := 2000) {{
    deadline := A_TickCount + timeoutMs
    while A_TickCount < deadline {{
        if IsAnyFileDialogActive() {{
            return true
        }}
        Sleep 80
    }}
    return false
}}

MarkCurrentDoneAndNext() {{
    global records, currentIndex

    records[currentIndex].done := true
    records[currentIndex].submitTime := FormatTime(, "yyyy-MM-dd HH:mm:ss")
    SaveProgress()
    Toast("源行 " records[currentIndex].sourceRow " 已标记为已填")

    if !MoveToPending(1) {{
        Toast("全部记录已处理完成，脚本即将退出")
        SetTimer(() => ExitApp(), -900)
    }}
}}

OpenPhotoPath() {{
    global records, currentIndex

    rawPath := records[currentIndex].photoPath
    fullPath := ResolvePhotoPath(rawPath)
    if fullPath = "" {{
        Toast("照片不存在或路径为空: " rawPath)
        return
    }}

    Run fullPath
}}

ShowCurrentRowValues() {{
    global records, currentIndex

    rec := records[currentIndex]
    content := "当前记录索引: " currentIndex "`n"
    content .= "Excel源行: " rec.sourceRow "`n`n"
    content .= "问题地址: " rec.address "`n"
    content .= "问题路段: " rec.section "`n"
    content .= "截止时间(小时数): " rec.deadlineHours "`n"
    content .= "问题类别: " rec.category "`n"
    content .= "问题描述: " rec.description "`n"
    content .= "上传照片路径: " rec.photoPath "`n"
    content .= "处置方式: " rec.disposal "`n"
    content .= "状态: " (rec.done ? "已填" : "待填")
    MsgBox content
}}

ResolvePhotoPath(rawPath) {{
    path := Trim(rawPath)
    if path = "" {{
        return ""
    }}

    fullPath := path
    if !FileExist(fullPath) {{
        fullPath := A_ScriptDir "\\" path
    }}

    if !FileExist(fullPath) {{
        return ""
    }}
    return fullPath
}}

SaveAndExit() {{
    SaveProgress()
    ExitApp
}}

SafeInt(value, defaultValue := 0) {{
    try {{
        return Integer(value)
    }} catch {{
        return defaultValue
    }}
}}

Toast(message) {{
    ToolTip message
    SetTimer(() => ToolTip(), -1200)
}}
'''
