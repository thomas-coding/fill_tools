from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


BASE_DIR = Path(__file__).resolve().parent
PROGRESS_TSV_PATH = BASE_DIR / "wechat_form_progress.tsv"
META_PATH = BASE_DIR / "wechat_source_meta.txt"
SOURCE_CANDIDATES = ["1.excel", "1.xlsx", "wechat_form_test.xlsx"]


def norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


def norm_header(value: object) -> str:
    return re.sub(r"\s+", "", norm(value))


def load_progress_map() -> Dict[int, Tuple[str, str]]:
    progress: Dict[int, Tuple[str, str]] = {}
    if not PROGRESS_TSV_PATH.exists():
        return progress

    lines = PROGRESS_TSV_PATH.read_text(encoding="utf-8").splitlines()
    for line in lines[1:]:
        if not line.strip():
            continue
        cols = line.split("\t")
        if len(cols) < 3:
            continue
        try:
            row_num = int(cols[0])
        except ValueError:
            continue
        progress[row_num] = (norm(cols[1]), norm(cols[2]))

    return progress


def find_source_workbook() -> Path:
    if META_PATH.exists():
        lines = META_PATH.read_text(encoding="utf-8").splitlines()
        for line in lines:
            if line.startswith("source="):
                p = Path(line.split("=", 1)[1])
                if p.exists():
                    return p

    for name in SOURCE_CANDIDATES:
        p = BASE_DIR / name
        if p.exists():
            return p

    xlsx_files = sorted(BASE_DIR.glob("*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]

    raise FileNotFoundError("未找到可回写的 Excel 文件")


def normalize_source_path(source_path: Path) -> Path:
    if source_path.suffix.lower() != ".excel":
        return source_path

    xlsx_path = source_path.with_suffix(".xlsx")
    if xlsx_path.exists():
        return xlsx_path

    xlsx_path.write_bytes(source_path.read_bytes())
    return xlsx_path


def infer_mode(ws: Worksheet) -> str:
    first_row_headers = {norm_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
    if "问题道路" in first_row_headers and "具体位置" in first_row_headers and "问题表述" in first_row_headers:
        return "patrol"
    return "simple"


def safe_set(ws: Worksheet, row_num: int, col_num: int, value: str) -> None:
    try:
        cell = ws.cell(row_num, col_num)
        setattr(cell, "value", value)
    except Exception:
        pass


def write_simple_mode(ws: Worksheet, progress_map: Dict[int, Tuple[str, str]]) -> int:
    if not norm(ws.cell(1, 8).value):
        safe_set(ws, 1, 8, "状态")
    if not norm(ws.cell(1, 9).value):
        safe_set(ws, 1, 9, "提交时间")

    updated = 0
    for row_num, (status, submit_time) in progress_map.items():
        safe_set(ws, row_num, 8, status)
        safe_set(ws, row_num, 9, submit_time)
        updated += 1
    return updated


def write_patrol_mode(ws: Worksheet, progress_map: Dict[int, Tuple[str, str]]) -> int:
    updated = 0
    for row_num, (status, submit_time) in progress_map.items():
        if row_num <= 1:
            continue

        header_row = row_num - 1
        safe_set(ws, header_row, 16, "填报状态")
        safe_set(ws, header_row, 17, "提交时间")
        safe_set(ws, row_num, 16, status)
        safe_set(ws, row_num, 17, submit_time)
        updated += 1
    return updated


def write_patrol_mode_com(source_path: Path, progress_map: Dict[int, Tuple[str, str]]) -> int:
    try:
        import win32com.client  # type: ignore
    except Exception:
        print("巡检模板回写需要 Excel COM，当前环境不可用，已跳过回写。")
        return 0

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    updated = 0
    try:
        wb = excel.Workbooks.Open(str(source_path))
        ws = wb.Worksheets(1)
        for row_num, (status, submit_time) in progress_map.items():
            if row_num <= 1:
                continue
            ws.Cells(row_num - 1, 16).Value = "填报状态"
            ws.Cells(row_num - 1, 17).Value = "提交时间"
            ws.Cells(row_num, 16).Value = status
            ws.Cells(row_num, 17).Value = submit_time
            updated += 1
        wb.Save()
        wb.Close(False)
        wb = None
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        excel.Quit()

    return updated


def main() -> int:
    try:
        source_path = normalize_source_path(find_source_workbook())
    except FileNotFoundError as err:
        print(str(err))
        return 1

    progress_map = load_progress_map()
    if not progress_map:
        print("未找到有效进度，不需要回写。")
        return 0

    try:
        wb = load_workbook(source_path)
    except PermissionError:
        print("Excel 正在被占用，已跳过状态回写。")
        return 0

    try:
        ws = wb.active
        assert ws is not None

        mode = infer_mode(ws)
        if mode == "patrol":
            wb.close()
            updated = write_patrol_mode_com(source_path, progress_map)
            print(f"已回写状态到 {source_path.name}，行数: {updated}")
            return 0

        updated = write_simple_mode(ws, progress_map)

        wb.save(source_path)
    except PermissionError:
        print("Excel 正在被占用，已跳过状态回写。")
        return 0
    finally:
        wb.close()

    print(f"已回写状态到 {source_path.name}，行数: {updated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
